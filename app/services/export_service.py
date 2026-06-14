from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


def build_reports_pdf(report_data: dict) -> BytesIO:
    """
    Build a PDF summary report using reportlab.
    """
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=20,
        leftMargin=20,
        topMargin=20,
        bottomMargin=20,
    )

    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph("Federal Polytechnic Ilaro - Institutional Report", styles["Title"]))
    story.append(Spacer(1, 12))
    story.append(Paragraph("Summary Metrics", styles["Heading2"]))
    story.append(Spacer(1, 8))

    summary_rows = [["Metric", "Value"]]
    for card in report_data.get("cards", []):
        summary_rows.append([str(card["label"]), str(card["value"])])

    summary_table = Table(summary_rows, colWidths=[220, 140])
    summary_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#10b981")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
                ("BACKGROUND", (0, 1), (-1, -1), colors.whitesmoke),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
                ("PADDING", (0, 0), (-1, -1), 8),
            ]
        )
    )
    story.append(summary_table)
    story.append(Spacer(1, 16))

    story.append(Paragraph("Submission Records", styles["Heading2"]))
    story.append(Spacer(1, 8))

    headers = [
        "Student",
        "Matric No",
        "Department",
        "Supervisor",
        "Project Title",
        "Stage",
        "Status",
        "Similarity %",
        "Risk",
        "Submitted At",
    ]
    rows = [headers]

    for row in report_data.get("rows", []):
        rows.append([
            str(row["student"]),
            str(row["matric_no"]),
            str(row["department"]),
            str(row["supervisor"]),
            str(row["project_title"]),
            str(row["stage"]),
            str(row["status"]),
            f'{row["similarity"]}%',
            str(row["risk"]),
            str(row["submitted_at"]),
        ])

    table = Table(rows, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f172a")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#cbd5e1")),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("BACKGROUND", (0, 1), (-1, -1), colors.white),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
                ("PADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    story.append(table)

    doc.build(story)
    buffer.seek(0)
    return buffer


def build_reports_xlsx(report_data: dict) -> BytesIO:
    """
    Build an Excel workbook using openpyxl.
    """
    buffer = BytesIO()
    wb = Workbook()

    # Summary sheet
    ws1 = wb.active
    ws1.title = "Summary"
    ws1["A1"] = "Federal Polytechnic Ilaro - Institutional Report"
    ws1["A1"].font = Font(bold=True, size=14)
    ws1["A3"] = "Metric"
    ws1["B3"] = "Value"
    ws1["A3"].font = Font(bold=True)
    ws1["B3"].font = Font(bold=True)
    ws1["A3"].fill = PatternFill("solid", fgColor="10B981")
    ws1["B3"].fill = PatternFill("solid", fgColor="10B981")
    ws1["A3"].font = Font(bold=True, color="FFFFFF")
    ws1["B3"].font = Font(bold=True, color="FFFFFF")

    row_idx = 4
    for card in report_data.get("cards", []):
        ws1[f"A{row_idx}"] = card["label"]
        ws1[f"B{row_idx}"] = card["value"]
        row_idx += 1

    ws1.column_dimensions["A"].width = 28
    ws1.column_dimensions["B"].width = 18

    # Reports sheet
    ws2 = wb.create_sheet("Reports")
    headers = [
        "Student",
        "Matric No",
        "Department",
        "Supervisor",
        "Project Title",
        "Stage",
        "Status",
        "Similarity %",
        "Risk",
        "Submitted At",
    ]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws2.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0F172A")
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(report_data.get("rows", []), start=2):
        ws2.cell(row=row_idx, column=1, value=row["student"])
        ws2.cell(row=row_idx, column=2, value=row["matric_no"])
        ws2.cell(row=row_idx, column=3, value=row["department"])
        ws2.cell(row=row_idx, column=4, value=row["supervisor"])
        ws2.cell(row=row_idx, column=5, value=row["project_title"])
        ws2.cell(row=row_idx, column=6, value=row["stage"])
        ws2.cell(row=row_idx, column=7, value=row["status"])
        ws2.cell(row=row_idx, column=8, value=row["similarity"])
        ws2.cell(row=row_idx, column=9, value=row["risk"])
        ws2.cell(row=row_idx, column=10, value=row["submitted_at"])

    widths = [24, 16, 20, 24, 28, 16, 16, 14, 12, 20]
    for i, width in enumerate(widths, start=1):
        ws2.column_dimensions[ws2.cell(row=1, column=i).column_letter].width = width

    wb.save(buffer)
    buffer.seek(0)
    return buffer